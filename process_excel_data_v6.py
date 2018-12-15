import re
import os
import openpyxl
import datetime
from dateutil.parser import parse

dur_from_proj = []
date_from_proj = []
pct_complete = []
date_of_calc = []
d_numeratior =[]
predicted_pct = []
date_datetime = []
predicted_pct = []
final_output = []
duration_total = 0
d_numeratior_sum = 0
total_pct_complete = 0
output_predict_pct = 0
last_date = datetime.datetime.min
todays_date = datetime.date.today()
first_date = datetime.datetime.max

#first_date = datetime.datetime(2025, 1, 21, 9, 54)

#TODO replace with a reasonable interface to select file
path = "C:\\Users\\dthompson\\Python\\Design change plan 12-12-2018_updated.xlsx"

# open the excel file
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object from the active attribute
#proj_sheet_obj = wb_obj.active # sets active sheet to proj_sheet_obj
proj_sheet_obj = wb_obj['Task_Data'] # sets sheet to Task Data 

KPIws = wb_obj.create_sheet("simpleKPI") # create a new sheet for kpi data
KPIws.sheet_properties.tabColor = "1072BA" # change tab color

wb_obj.save(path) 
wb_obj.active = KPIws
KPI_sheet_obj = wb_obj.active # this might be redundant w KPIws
 
# some other functions that are not needed 
# print(wb_obj.sheetnames)
# print(wb_obj.active)
# print(wb_obj._active_sheet_index)

# Cell object is created by using proj sheet object
# 's cell() method.
cell_obj = proj_sheet_obj.cell(row = 1, column = 1)

max_col = proj_sheet_obj.max_column 
max_row = proj_sheet_obj.max_row

# read in all data 
for i in range(2, max_row + 1): 
    duration_temp = proj_sheet_obj.cell(row = i, column = 5) 
    date_temp = proj_sheet_obj.cell(row = i, column = 6) 
    pct_complete_temp = proj_sheet_obj.cell(row = i, column = 8) 

    dur_from_proj.append(duration_temp)
    date_from_proj.append(date_temp)
    pct_complete.append(pct_complete_temp)
    pct_complete_temp
    
for i in range(0, max_row-1):
    dur_1 = str(dur_from_proj[i].value)
    if dur_1 != 'None':
        dur_1 = float(dur_1[0:dur_1.index(" day")]) # remove "days" and convert
    else:
        dur_1 = 0.0
    dur_from_proj[i] = dur_1 # load into list as a float
    duration_total = duration_total + dur_from_proj[i]
    date_1 = date_from_proj[i].value  # start date in readable format
    pct_complete[i] = float(pct_complete[i].value) # convert to float
    total_pct_complete =  total_pct_complete + dur_from_proj[i] * pct_complete[i]     
    
    if date_1 != None:
        date_datetime.append(parse(date_1)) # convert to datetime - matched
    else:
        date_datetime.append(datetime.datetime(2025, 1, 21, 9, 54) ) # a long way in the future
#   find earliest date in dataset
    if date_datetime[i] < first_date: 
        first_date = date_datetime[i] # calc earliest date 
    if date_datetime[i] > last_date: # calc number of weeks in proj
        last_date = date_datetime[i] 
        proj_length = ( (last_date - first_date).days ) / 7
        proj_length = int(proj_length) + 1

print("date_from_proj,", "planned %,", "actual %")
#TODO replace max in loop with max date_from_proj +7 days
for i in range(0,proj_length): # calc expected progress by date of calc then inc by i weeks
    date_of_calc = first_date + datetime.timedelta(days=7*i) # increment by a week - matched
    if date_1 != None:
        for j in range(0,max_row-1): # looping in j for each row in table
            d_temp = date_of_calc - date_datetime[j]
            d_temp = float(d_temp.days)

            if d_temp > dur_from_proj[j]:
                d_numeratior.append(dur_from_proj[j])
            elif d_temp < 0:
                d_numeratior.append(float(0.0))
            else:
                d_numeratior.append(d_temp)

            d_numeratior_sum = d_numeratior_sum + d_numeratior[j]
    
        output_predict_pct = d_numeratior_sum
        d_numeratior_sum = 0 # reset for next summing loop

        output_predict_pct = output_predict_pct / duration_total
        final_output.append([date_of_calc,output_predict_pct,total_pct_complete])
        d_numeratior = [] # reset for next loop
        predicted_pct = [] # reset for next loop
        output_predict_pct = 0 # reset for next loop

total_pct_complete = total_pct_complete / duration_total
# write data to terminal in readable format
for i in range(0,proj_length): # need to determine what replaces the 4
    print("{0:%Y-%m-%d},{1:8.2f},{2:8.2f}".format(final_output[i][0], final_output[i][1],total_pct_complete ) )

# write data in KPI format to a new sheet in the project output file 
KPI_sheet_obj.cell(row=1, column=1).value = "Email"
KPI_sheet_obj.cell(row=1, column=2).value = "KPID"
KPI_sheet_obj.cell(row=1, column=3).value = "KPI Name"
KPI_sheet_obj.cell(row=1, column=4).value = "Date"
KPI_sheet_obj.cell(row=1, column=5).value = "Target"
KPI_sheet_obj.cell(row=1, column=6).value = "Actual"
KPI_sheet_obj.cell(row=1, column=7).value = "Notes"

for i in range(0,proj_length):
    KPI_sheet_obj.cell(row=i+2, column=1).value = "David.Thompson@livanova.com"
    KPI_sheet_obj.cell(row=i+2, column=2).value = "309"
    KPI_sheet_obj.cell(row=i+2, column=3).value = "ImThera FW+SW rewrite"
    KPI_sheet_obj.cell(row=i+2, column=4).value = final_output[i][0]
    KPI_sheet_obj.cell(row=i+2, column=5).value = final_output[i][1]
    KPI_sheet_obj.cell(row=i+2, column=6).value = total_pct_complete

# TODO change to a new file by appending suffix to path
# so that simpleKPI can read it easily
# TODO fix format of date so same as terminal. KPI may not like the time included
wb_obj.save(path) 
